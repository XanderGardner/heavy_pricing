import os
from os.path import exists
import math
import sys
import openpyxl as pyxl
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from threading import Thread
import time

# INPUT/ CONSTANTS
SAVE_EVERY = 10 # save excel after every SAVE_EVERY number of elements scraped
MAX_NUM_TO_SCRAPE = 999999 # max number of elements to scrape (set high to scrape everything)
MAX_INCREASE = 2 # max amount allowed for given auction value to increase before result is not used
MAX_DECREASE = 0.4 # max amount allowed for given auction value to decrease before result is not used
HEADLESS = False # if running with chrome browser showing (more results when false, but takes longer)

MAX_THREADS = 10 # max reasonable number of threads as each thread has a chomium driver
OFFSET = 2 # excel input data is offset by 2: 1 for 0 indexing and 1 for a row of titles
OFFSET_ROWS = 2 # excel input data has 2 extra rows: first is headers, last row is totaled info

# imports data from 'Equipment New List.xlsx' and returns as a dictionary
def getExcelValues():

    wb = pyxl.load_workbook('Equipment New List.xlsx')
    ws = wb.active
    n = ws.max_row - OFFSET_ROWS
    if n > MAX_NUM_TO_SCRAPE:
        n = MAX_NUM_TO_SCRAPE

    a1 = [None] * n
    a2 = [None] * n
    a3 = [None] * n
    a4 = [None] * n
    a5 = [None] * n
    a6 = [None] * n
    a7 = [None] * n
    a8 = [None] * n
    a9 = [None] * n
    a10 = [None] * n
    a11 = [None] * n
    a12 = [None] * n
    a13 = [None] * n
    a14 = [None] * n
    a15 = [None] * n
    a16 = [None] * n
    a17 = [None] * n
    a18 = [None] * n
    a19 = [None] * n
    a20 = [None] * n
    a21 = [None] * n
    a22 = [None] * n
    a23 = [None] * n
    a24 = [None] * n
    a25 = [None] * n
    a26 = [None] * n
    a27 = [None] * n
    a28 = [None] * n
    a29 = [None] * n
    a30 = [None] * n
    a31 = [None] * n
    a32 = [None] * n
    a33 = [None] * n
    a34 = [None] * n
    a35 = [None] * n
    a36 = [None] * n
    a37 = [None] * n
    a38 = [None] * n
    a39 = [None] * n
    
    i = 0
    while i != n:
        a1[i] = ws[f'A{i+OFFSET}'].value
        a2[i] = ws[f'B{i+OFFSET}'].value
        a3[i] = ws[f'C{i+OFFSET}'].value
        a4[i] = ws[f'D{i+OFFSET}'].value
        a5[i] = ws[f'E{i+OFFSET}'].value
        a6[i] = ws[f'F{i+OFFSET}'].value
        a7[i] = ws[f'G{i+OFFSET}'].value
        a8[i] = ws[f'H{i+OFFSET}'].value
        a9[i] = ws[f'I{i+OFFSET}'].value
        a10[i] = ws[f'J{i+OFFSET}'].value
        a11[i] = ws[f'K{i+OFFSET}'].value
        a12[i] = ws[f'L{i+OFFSET}'].value
        a13[i] = ws[f'M{i+OFFSET}'].value
        a14[i] = ws[f'N{i+OFFSET}'].value
        a15[i] = ws[f'O{i+OFFSET}'].value
        a16[i] = ws[f'P{i+OFFSET}'].value
        a17[i] = ws[f'Q{i+OFFSET}'].value
        a18[i] = ws[f'R{i+OFFSET}'].value
        a19[i] = ws[f'S{i+OFFSET}'].value
        a20[i] = ws[f'T{i+OFFSET}'].value
        a21[i] = ws[f'U{i+OFFSET}'].value
        a22[i] = ws[f'V{i+OFFSET}'].value
        a23[i] = ws[f'W{i+OFFSET}'].value
        a24[i] = ws[f'X{i+OFFSET}'].value
        a25[i] = ws[f'Y{i+OFFSET}'].value
        a26[i] = ws[f'Z{i+OFFSET}'].value
        a27[i] = ws[f'AA{i+OFFSET}'].value
        a28[i] = ws[f'AB{i+OFFSET}'].value
        a29[i] = ws[f'AC{i+OFFSET}'].value
        a30[i] = ws[f'AD{i+OFFSET}'].value
        a31[i] = ws[f'AE{i+OFFSET}'].value
        a32[i] = ws[f'AF{i+OFFSET}'].value
        a33[i] = ws[f'AG{i+OFFSET}'].value
        a34[i] = ws[f'AH{i+OFFSET}'].value
        a35[i] = ws[f'AI{i+OFFSET}'].value
        a36[i] = ws[f'AJ{i+OFFSET}'].value
        a37[i] = ws[f'AK{i+OFFSET}'].value
        a38[i] = ws[f'AL{i+OFFSET}'].value
        a39[i] = ws[f'AM{i+OFFSET}'].value
        
        i += 1
    
    data = {
        'Emco' : a1,
        'Equipment' : a2,
        'Description' : a3,
        'VINNumber' : a4,
        'Manufacturer' : a5,
        'Model' : a6,
        'ModelYr' : a7,
        'OdoReading' : a8,
        'OdoDate' : a9,
        'HourReading' : a10,
        'HourData' : a11,
        'Location' : a12,
        'Complete' : a13,
        'Auction Value' : a14,
        'Market Value' : a15,
        'Asking Value' : a16,

        'Market Value Found' : a17,
        'Auction Value Found' : a18,
        'Auction Value Link' : a19,
        'Asking Value Found' : a20,
        'Asking Value Link' : a21,
        'gmvf1' : a22,
        'gmvl1' : a23,
        'gmvf2' : a24,
        'gmvl2' : a25,
        'gmvf3' : a26,
        'gmvl3' : a27,
        'gmvf4' : a28,
        'gmvl4' : a29,
        'gmvf5' : a30,
        'gmvl5' : a31,
        'gmvf6' : a32,
        'gmvl6' : a33,
        'gmvf7' : a34,
        'gmvl7' : a35,
        'gmvf8' : a36,
        'gmvl8' : a37,
        'gmvf9' : a38,
        'gmvl9' : a39,
    }

    return data

# create dict from previously found scraped data given in data
def getDict(data):
    def int_none(val):
        if val:
            return int(val)
        return None
    dict = {}
    dict['Market Value Found'] = list(map(int_none, data['Market Value Found']))
    dict['Auction Value Found'] = list(map(int_none, data['Auction Value Found']))
    dict['Auction Value Link'] = data['Auction Value Link'][:]
    dict['Asking Value Found'] = list(map(int_none, data['Asking Value Found']))
    dict['Asking Value Link'] = data['Asking Value Link'][:]
    for i in range(1, 10):
        dict[f'gmvf{i}'] = list(map(int_none, data[f'gmvf{i}']))
        dict[f'gmvl{i}'] = data[f'gmvl{i}'][:]
    
    return dict

# create search term for each item in data and return as array of search term strings
def get_search_terms(data):
    n = len(data['Emco'])
    search_terms = [None] * n
    for i in range(n):
        terms = []
        if data['Manufacturer'][i]:
            terms.append(data['Manufacturer'][i])
        if data['Model'][i]:
            terms.append(data['Model'][i])
        if data['ModelYr'][i]:
            terms.append(data['ModelYr'][i])
        search_term = ' '.join(str(term) for term in terms)
        if len(search_term) <= 8 and data['Description'][i]:
            search_terms[i] = f"{search_term} {data['Description'][i]}"
        else:
            search_terms[i] = search_term
    return search_terms

# create multi-round search term for each item in data and return as array of array of search term strings
def get_adv_search_terms(data):
    n = len(data['Emco'])
    adv_search_terms = [None] * n
    # find many ways of searching for each item
    for i in range(n):
        # get string values for the item
        manufacturer = str(data['Manufacturer'][i]) if data['Manufacturer'][i] else None
        model = str(data['Model'][i]) if data['Model'][i] else None
        description = str(data['Description'][i]) if data['Description'][i] else None
        model_year = str(data['ModelYr'][i]) if data['ModelYr'][i] else None

        # create search terms
        curr_terms = []
        if description and model and model_year:
            curr_terms.append(' '.join([description, model, model_year, "used price"]))
        if description and model:
            curr_terms.append(' '.join([description, model, "used price"]))
        if description and manufacturer:
            curr_terms.append(' '.join([description, manufacturer, "used price"]))
        if manufacturer and model and model_year:
            curr_terms.append(' '.join([manufacturer, model, model_year, "used price"]))
        adv_search_terms[i] = curr_terms
        
    return adv_search_terms

# scrape data from https://usedequipmentguide.com/ given a list of search terms
# in dict, updates dict with new found values and links with keys 'Asking Value Link' and 'Asking Value Found', and
# saves results to 'Equipment New List.xlsx' as it searches
def scrapeAskingValues(dict):
    # constants and variables
    search_terms = dict['Search Terms']
    n = len(search_terms)
    avf = dict['Asking Value Found'] # asking values found
    avl = dict['Asking Value Link'] # asking value links
    
    # parse data: convert given dollar string to an int
    def parseDollarValue(money_str):
        str = money_str[1:] # remove dollar sign
        value = 0
        i = len(str) - 1
        multiplier = 1
        while i >= 0:
            if (len(str) - i) % 4 == 0:
                i -= 1
                continue
            else:
                value += multiplier * int(str[i])
                multiplier *= 10
                i -= 1
        return value

    # nested function for threaded scraping
    def scrape_task(index):
        chrome_options = Options()
        if HEADLESS:
            chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        driver.get(f"https://usedequipmentguide.com/listings?query={search_terms[index]}")
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "span.Span-hup779-0.sc-16afded-0.kzgLyd")
            )) # will only wait for first result
            time.sleep(1.5) # testing shows that an extra 1.5 seconds allows all results to finish
            elements = driver.find_elements(by=By.CSS_SELECTOR, value="span.Span-hup779-0.sc-16afded-0.kzgLyd")
            
            # pick most relavent result
            i = 0
            while i < len(elements):
                if elements[i].text != "AUCTION" and elements[i].text != "Price Unavailable" and parseDollarValue(elements[i].text) > 999:
                    avf[index] = parseDollarValue(elements[i].text)
                    avl[index] = f"https://usedequipmentguide.com/listings?query={search_terms[index]}"
                    break
                i += 1
        
        finally:
            driver.close()
            driver.quit()
    
    i = n - 1
    # start scraping where it was last stopped
    while not (i == -1 or avf[i]):
        i -= 1
    i += 1  
    i = i // SAVE_EVERY * SAVE_EVERY # corner case: round down to nearest SAVE_EVERY
    while i < n:
        # run next set of threads
        threads = [None] * MAX_THREADS
        ti = 0
        while ti < MAX_THREADS and i < n:
            threads[ti] = Thread(target=scrape_task, args=(i,))
            threads[ti].start()
            i += 1
            ti += 1
        for j in range(ti):
            threads[j].join()

        # occasionally save what is found
        if i % SAVE_EVERY == 0:
            row_start = i-SAVE_EVERY
            temp_dict = {
                'Asking Value Found' : avf[i-SAVE_EVERY:i],
                'Asking Value Link' : avl[i-SAVE_EVERY:i]
            }
            tempSetExcel(temp_dict, row_start)

    # save final results
    row_start = i-SAVE_EVERY
    # corner case: SAVE_EVERY is large and row start becomes negative
    if row_start < 0:
        row_start = 0
    temp_dict = {
        'Asking Value Found' : avf[i-SAVE_EVERY:i],
        'Asking Value Link' : avl[i-SAVE_EVERY:i]
    }
    tempSetExcel(temp_dict, row_start)

    return 1

# scrape data from ebay's trucks and cars site given a list of search terms
# in dict, updates dict with new found values and links with keys 'Auction Value Link' and 'Auction Value Found', and
# saves results to 'Equipment New List.xlsx' as it searches
def scrapeAuctionValues(dict):
    # constants and variables
    search_terms = dict['Search Terms']
    n = len(search_terms)
    avf = dict['Auction Value Found'] # auction values found
    avl = dict['Auction Value Link'] # auction value links
    
    # parse ebay data: convert given dollar string to an int
    def parseDollarValue(money_str):
        str = money_str[1:-3] # remove dollar sign and pennies
        value = 0.01 * int(money_str[-2:]) # value in the pennies
        i = len(str) - 1
        multiplier = 1
        while i >= 0:
            if (len(str) - i) % 4 == 0:
                i -= 1
                continue
            else:
                value += multiplier * int(str[i])
                multiplier *= 10
                i -= 1
        return value

    # nested function for threaded scraping
    def scrape_task(index):
        chrome_options = Options()
        if HEADLESS:
            chrome_options.add_argument("--headless")  
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        driver.get(f"https://www.ebay.com/sch/i.html?_from=R40&_nkw={search_terms[index]}&_sacat=6001&rt=nc&LH_Sold=1&LH_Complete=1")
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "li.s-item.s-item__pl-on-bottom")
            )) # will only wait for first result
            time.sleep(1.5) # testing shows that an extra 1.5 seconds allows all results to finish
            
            elements = driver.find_elements(by=By.CSS_SELECTOR, value="span.POSITIVE")
            
            # pick only relevent trucks and cars data
            if len(elements) >= 1:
                results_num_el = driver.find_elements(by=By.CSS_SELECTOR, value="span.section-notice__main")
                main_results = driver.find_elements(by=By.CSS_SELECTOR, value="h1.srp-controls__count-heading")
                main_num = main_results[0].text[0]
                str_dollar_value = str(elements[1].text)  # 2nd element is the 1st most relevent price
                dollar_value = parseDollarValue(str_dollar_value)
                if (len(results_num_el) == 0 or results_num_el[0].text[0] != "0") and dollar_value > 999 and main_num != "0":
                    avf[index] = dollar_value
                    avl[index] = f"https://www.ebay.com/sch/i.html?_from=R40&_nkw={search_terms[index]}&_sacat=6001&rt=nc&LH_Sold=1&LH_Complete=1"
                
        finally:
            driver.close()
            driver.quit()
    
    i = n - 1
    # start scraping where it was last stopped
    while not (i == -1 or avf[i]):
        i -= 1
    i += 1  
    i = i // SAVE_EVERY * SAVE_EVERY # corner case: round down to nearest SAVE_EVERY
    while i < n:
        # run next set of threads
        threads = [None] * MAX_THREADS
        ti = 0
        while ti < MAX_THREADS and i < n:
            threads[ti] = Thread(target=scrape_task, args=(i,))
            threads[ti].start()
            i += 1
            ti += 1
        for j in range(ti):
            threads[j].join()

        # occasionally save what is found
        if i % SAVE_EVERY == 0:
            row_start = i-SAVE_EVERY
            temp_dict = {
                'Auction Value Found' : avf[i-SAVE_EVERY:i],
                'Auction Value Link' : avl[i-SAVE_EVERY:i]
            }
            tempSetExcel(temp_dict, row_start)

    # save final results
    row_start = i-SAVE_EVERY
    # corner case: SAVE_EVERY is large and row start becomes negative
    if row_start < 0:
        row_start = 0
    temp_dict = {
        'Auction Value Found' : avf[i-SAVE_EVERY:i],
        'Auction Value Link' : avl[i-SAVE_EVERY:i]
    }
    tempSetExcel(temp_dict, row_start)

    return 1

# scrape data from google's site given a list of search terms
# in dict, updates dict with new found values and links with keys 'General Market Value Link' and 'General Market Value Found', and
# saves results to 'Equipment New List.xlsx' as it searches
def scrapeGeneralMarketValues(dict):
    # constants and variables
    search_terms = dict['Search Terms']
    adv_search_terms = dict['Advanced Search Terms']
    n = len(search_terms)
    gmvf1 = dict['gmvf1']
    gmvl1 = dict['gmvl1']
    gmvf2 = dict['gmvf2']
    gmvl2 = dict['gmvl2']
    gmvf3 = dict['gmvf3']
    gmvl3 = dict['gmvl3']
    gmvf4 = dict['gmvf4']
    gmvl4 = dict['gmvl4']
    gmvf5 = dict['gmvf5']
    gmvl5 = dict['gmvl5']
    gmvf6 = dict['gmvf6']
    gmvl6 = dict['gmvl6']
    gmvf7 = dict['gmvf7']
    gmvl7 = dict['gmvl7']
    gmvf8 = dict['gmvf8']
    gmvl8 = dict['gmvl8']
    gmvf9 = dict['gmvf9']
    gmvl9 = dict['gmvl9']
    
    # return average dollar seen or -1 if not valid
    def parseDollarValue(dom_text):
        arr_dom_text = dom_text.split()
        dollar_strs = []
        for text in arr_dom_text:
            if text[0] == "$":
                dollar_strs.append(text[1:])
        
        acceptable_chars = {"0", "1", "2", "3", "4", "5", "6", "7", "8", "9", ","}
        dollar_values = []
        for dollar_str in dollar_strs:
            i = 0
            while i < len(dollar_str) and dollar_str[i] in acceptable_chars:
                i += 1
            i -= 1
            # extract numerical value
            multiplier = 1
            value = 0
            while i >= 0:
                if dollar_str[i] == ",":
                    i -= 1
                    continue
                else:
                    value += multiplier * int(dollar_str[i])
                    multiplier *= 10
                    i -= 1
            dollar_values.append(value)
        dollar_values.sort()
        num_values = len(dollar_values)
        if num_values == 0:
            return -1
        elif num_values <= 5:
            return dollar_values[math.floor(num_values / 2)] # return median
        else:
            # remove outliers
            min = round(0.3 * num_values) 
            max = round(0.8 * num_values)

            # average the remaining
            return sum(dollar_values[min:max]) / len(dollar_values[min:max])

    # nested function for threaded scraping
    def scrape_task(index):
        def scrape_engine(address):
            price = None
            chrome_options = Options()
            if HEADLESS:
                chrome_options.add_argument("--headless")
            driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
            driver.get(address)
            
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.TAG_NAME, "body")
                ))
                time.sleep(1.0) # testing shows that an extra second allows all results to load
                
                main_element = driver.find_element(by=By.TAG_NAME, value="body")
                dom_text = main_element.text
                dollar_value = parseDollarValue(dom_text)
                if dollar_value > 100:
                    engine_price_decrease = 0.88
                    price = engine_price_decrease * dollar_value
            finally:
                driver.close()
                driver.quit()
                return price

        # search all search engines
        curr_search_terms = [search_terms[index] + " used price"] + adv_search_terms[index]
        price = scrape_engine(f"https://www.bing.com/search?q={curr_search_terms[0]}")
        if price:
            gmvf1[index] = price
            gmvl1[index] = f"https://www.bing.com/search?q={curr_search_terms[0]}"
        price = scrape_engine(f"https://swisscows.com/web?query={curr_search_terms[0]}")
        if price:
            gmvf2[index] = price
            gmvl2[index] = f"https://swisscows.com/web?query={curr_search_terms[0]}"
        price = scrape_engine(f"https://duckduckgo.com/?q={curr_search_terms[0]}&ia=web")
        if price:
            gmvf3[index] = price
            gmvl3[index] = f"https://duckduckgo.com/?q={curr_search_terms[0]}&ia=web"
        price = scrape_engine(f"https://gibiru.com/results.html?q={curr_search_terms[0]}")
        if price:
            gmvf4[index] = price
            gmvl4[index] = f"https://gibiru.com/results.html?q={curr_search_terms[0]}"
        price = scrape_engine(f"https://search.givewater.com/serp?q={curr_search_terms[0]}")
        if price:
            gmvf5[index] = price
            gmvl5[index] = f"https://search.givewater.com/serp?q={curr_search_terms[0]}"
        price = scrape_engine(f"https://ekoru.org/?q={curr_search_terms[0]}")
        if price:
            gmvf6[index] = price
            gmvl6[index] = f"https://ekoru.org/?q={curr_search_terms[0]}"
        price = scrape_engine(f"https://www.ecosia.org/search?method=index&q={curr_search_terms[0]}")
        if price:
            gmvf7[index] = price
            gmvl7[index] = f"https://www.ecosia.org/search?method=index&q={curr_search_terms[0]}"
        price = scrape_engine(f"https://search.brave.com/search?q={curr_search_terms[0]}&source=web")
        if price:
            gmvf8[index] = price
            gmvl8[index] = f"https://search.brave.com/search?q={curr_search_terms[0]}&source=web"
        
        # advanced iterations with different search terms until out of options or found price on google
        while not (gmvf9[index] or len(curr_search_terms) == 0 or gmvf8[index] or gmvf7[index] or gmvf6[index] or gmvf5[index] or gmvf4[index] or gmvf3[index] or gmvf2[index] or gmvf1[index]):
            chrome_options = Options()
            if HEADLESS:
                chrome_options.add_argument("--headless") 
            search_term = curr_search_terms.pop(0)
            driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
            driver.get(f"https://www.google.com/search?q={search_term}")
            
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.ID, "main")
                ))
                time.sleep(1.0) # testing shows that an extra second allows all results to finish
                
                main_element = driver.find_element(by=By.ID, value="main")
                dom_text = main_element.text
                dollar_value = parseDollarValue(dom_text)
                if dollar_value > 100:
                    google_price_decrease = 0.88
                    gmvf9[index] = google_price_decrease * dollar_value
                    gmvl9[index] = f"https://www.google.com/search?q={search_term}"
            finally:
                driver.close()
                driver.quit()
    
    i = n - 1
    # start scraping where it was last stopped
    while not (i == -1 or gmvf1[i] or gmvf2[i] or gmvf3[i] or gmvf4[i] or gmvf5[i] or gmvf6[i] or gmvf7[i] or gmvf8[i] or gmvf9[i]):
        i -= 1
    i += 1  
    i = i // SAVE_EVERY * SAVE_EVERY # corner case: round down to nearest SAVE_EVERY
    while i < n:
        # run next set of threads
        threads = [None] * MAX_THREADS
        ti = 0
        while ti < MAX_THREADS and i < n:
            threads[ti] = Thread(target=scrape_task, args=(i,))
            threads[ti].start()
            i += 1
            ti += 1
        for j in range(ti):
            threads[j].join()

        # occasionally save what is found
        if i % SAVE_EVERY == 0:
            row_start = i-SAVE_EVERY
            temp_dict = {
                'gmvf1' : gmvf1[i-SAVE_EVERY:i],
                'gmvl1' : gmvl1[i-SAVE_EVERY:i],
                'gmvf2' : gmvf2[i-SAVE_EVERY:i],
                'gmvl2' : gmvl2[i-SAVE_EVERY:i],
                'gmvf3' : gmvf3[i-SAVE_EVERY:i],
                'gmvl3' : gmvl3[i-SAVE_EVERY:i],
                'gmvf4' : gmvf4[i-SAVE_EVERY:i],
                'gmvl4' : gmvl4[i-SAVE_EVERY:i],
                'gmvf5' : gmvf5[i-SAVE_EVERY:i],
                'gmvl5' : gmvl5[i-SAVE_EVERY:i],
                'gmvf6' : gmvf6[i-SAVE_EVERY:i],
                'gmvl6' : gmvl6[i-SAVE_EVERY:i],
                'gmvf7' : gmvf7[i-SAVE_EVERY:i],
                'gmvl7' : gmvl7[i-SAVE_EVERY:i],
                'gmvf8' : gmvf8[i-SAVE_EVERY:i],
                'gmvl8' : gmvl8[i-SAVE_EVERY:i],
                'gmvf9' : gmvf9[i-SAVE_EVERY:i],
                'gmvl9' : gmvl9[i-SAVE_EVERY:i]
            }
            tempSetExcel(temp_dict, row_start)

    # save final results
    row_start = i-SAVE_EVERY
    # corner case: SAVE_EVERY is large and row start becomes negative
    if row_start < 0:
        row_start = 0
    temp_dict = {
        'gmvf1' : gmvf1[i-SAVE_EVERY:i],
        'gmvl1' : gmvl1[i-SAVE_EVERY:i],
        'gmvf2' : gmvf2[i-SAVE_EVERY:i],
        'gmvl2' : gmvl2[i-SAVE_EVERY:i],
        'gmvf3' : gmvf3[i-SAVE_EVERY:i],
        'gmvl3' : gmvl3[i-SAVE_EVERY:i],
        'gmvf4' : gmvf4[i-SAVE_EVERY:i],
        'gmvl4' : gmvl4[i-SAVE_EVERY:i],
        'gmvf5' : gmvf5[i-SAVE_EVERY:i],
        'gmvl5' : gmvl5[i-SAVE_EVERY:i],
        'gmvf6' : gmvf6[i-SAVE_EVERY:i],
        'gmvl6' : gmvl6[i-SAVE_EVERY:i],
        'gmvf7' : gmvf7[i-SAVE_EVERY:i],
        'gmvl7' : gmvl7[i-SAVE_EVERY:i],
        'gmvf8' : gmvf8[i-SAVE_EVERY:i],
        'gmvl8' : gmvl8[i-SAVE_EVERY:i],
        'gmvf9' : gmvf9[i-SAVE_EVERY:i],
        'gmvl9' : gmvl9[i-SAVE_EVERY:i]
    }
    tempSetExcel(temp_dict, row_start)

    # set found prices and return
    return 1

# adds validated market values to dict with the key "Market Value Found" given the dict has the key-value 
# pairs for "Auction Value Found" and "General Market Value Found" and "Asking Value Found" 
# uses previous market value found in data to validate found prices according to MAX_INCREASE and MAX_DECREASE
# returns the number of successful market values found
def setMarketValues(dict, data):
    n = len(data["Auction Value"])

    auvf = dict["Auction Value Found"]
    auvf_link = dict["Auction Value Link"]
    asvf = dict["Asking Value Found"]
    asvf_link = dict["Asking Value Link"]
    gmvf1 = dict['gmvf1']
    gmvl1 = dict['gmvl1']
    gmvf2 = dict['gmvf2']
    gmvl2 = dict['gmvl2']
    gmvf3 = dict['gmvf3']
    gmvl3 = dict['gmvl3']
    gmvf4 = dict['gmvf4']
    gmvl4 = dict['gmvl4']
    gmvf5 = dict['gmvf5']
    gmvl5 = dict['gmvl5']
    gmvf6 = dict['gmvf6']
    gmvl6 = dict['gmvl6']
    gmvf7 = dict['gmvf7']
    gmvl7 = dict['gmvl7']
    gmvf8 = dict['gmvf8']
    gmvl8 = dict['gmvl8']
    gmvf9 = dict['gmvf9']
    gmvl9 = dict['gmvl9']
    auv = data["Auction Value"]
    mv  = data["Market Value"]
    asv = data["Asking Value"]

    market_values = [None] * n
    successes = 0
    for i in range(n):
        # remove found items that have changed too much from previous research
        if auvf[i] and auv[i] and (auvf[i] > auv[i] * MAX_INCREASE or auvf[i] < auv[i] * MAX_DECREASE):
            auvf[i] = None
            auvf_link[i] = None
        if asvf[i] and asv[i] and (asvf[i] > asv[i] * MAX_INCREASE or asvf[i] < asv[i] * MAX_DECREASE):
            asvf[i] = None
            asvf_link[i] = None
        if gmvf1[i] and mv[i] and (gmvf1[i] > mv[i] * MAX_INCREASE or gmvf1[i] < mv[i] * MAX_DECREASE):
            gmvf1[i] = None
            gmvl1[i] = None
        if gmvf2[i] and mv[i] and (gmvf2[i] > mv[i] * MAX_INCREASE or gmvf2[i] < mv[i] * MAX_DECREASE):
            gmvf2[i] = None
            gmvl2[i] = None
        if gmvf3[i] and mv[i] and (gmvf3[i] > mv[i] * MAX_INCREASE or gmvf3[i] < mv[i] * MAX_DECREASE):
            gmvf3[i] = None
            gmvl3[i] = None
        if gmvf4[i] and mv[i] and (gmvf4[i] > mv[i] * MAX_INCREASE or gmvf4[i] < mv[i] * MAX_DECREASE):
            gmvf4[i] = None
            gmvl4[i] = None
        if gmvf5[i] and mv[i] and (gmvf5[i] > mv[i] * MAX_INCREASE or gmvf5[i] < mv[i] * MAX_DECREASE):
            gmvf5[i] = None
            gmvl5[i] = None
        if gmvf6[i] and mv[i] and (gmvf6[i] > mv[i] * MAX_INCREASE or gmvf6[i] < mv[i] * MAX_DECREASE):
            gmvf6[i] = None
            gmvl6[i] = None
        if gmvf7[i] and mv[i] and (gmvf7[i] > mv[i] * MAX_INCREASE or gmvf7[i] < mv[i] * MAX_DECREASE):
            gmvf7[i] = None
            gmvl7[i] = None
        if gmvf8[i] and mv[i] and (gmvf8[i] > mv[i] * MAX_INCREASE or gmvf8[i] < mv[i] * MAX_DECREASE):
            gmvf8[i] = None
            gmvl8[i] = None
        if gmvf9[i] and mv[i] and (gmvf9[i] > mv[i] * MAX_INCREASE or gmvf9[i] < mv[i] * MAX_DECREASE):
            gmvf9[i] = None
            gmvl9[i] = None
        
        # examine how previous research shows getting market value from asking values
        decrease_asking = 0.9
        if mv[i] and asv[i]:
            decrease_asking = 1 + (mv[i] - asv[i]) / asv[i]
        if decrease_asking <= 0 or decrease_asking >= 1:
            decrease_asking = 0.9

        # find new market value
        values_to_avg = []
        if auvf[i]:
            values_to_avg.append(auvf[i])
        if asvf[i]:
            values_to_avg.append(asvf[i] * decrease_asking)
        if gmvf1[i]:
            values_to_avg.append(gmvf1[i])
        if gmvf2[i]:
            values_to_avg.append(gmvf2[i])
        if gmvf3[i]:
            values_to_avg.append(gmvf3[i])
        if gmvf4[i]:
            values_to_avg.append(gmvf4[i])
        if gmvf5[i]:
            values_to_avg.append(gmvf5[i])
        if gmvf6[i]:
            values_to_avg.append(gmvf6[i])
        if gmvf7[i]:
            values_to_avg.append(gmvf7[i])
        if gmvf8[i]:
            values_to_avg.append(gmvf8[i])
        if gmvf9[i]:
            values_to_avg.append(gmvf9[i])
        
        
        # set found market value
        if len(values_to_avg) > 0:
            market_values[i] = sum(values_to_avg) / len(values_to_avg)
            successes += 1
    
    dict["Market Value Found"] = market_values
    return successes

# arr_values is 2d array. Each item is an array representing data for a column 
# arr_col_strs is array of strings. arr_col_strs at index i is the col for arr_values at i
# tempSetExcel will set the values in the respective columns in 'Equipment New List.xlsx',
# starting at the row row_start
def tempSetExcel(dict, row_start):
    wb = pyxl.load_workbook('Equipment New List.xlsx')
    ws = wb.active

    # set market value found
    if 'Market Value Found' in dict:
        row = row_start
        for val in dict['Market Value Found']:
            ws[f'Q{row + OFFSET}'] = val
            row += 1
    
    # set auction values if they are given
    if 'Auction Value Found' in dict:
        row = row_start
        for val in dict['Auction Value Found']:
            ws[f'R{row + OFFSET}'] = val
            row += 1
    
    # set auction value links if they are given
    if 'Auction Value Link' in dict:
        row = row_start
        for val in dict['Auction Value Link']:
            ws[f'S{row + OFFSET}'] = val
            row += 1

    # set asking values if they are given
    if 'Asking Value Found' in dict:
        row = row_start
        for val in dict['Asking Value Found']:
            ws[f'T{row + OFFSET}'] = val
            row += 1

    # set asking value links if they are given
    if 'Asking Value Link' in dict:
        row = row_start
        for val in dict['Asking Value Link']:
            ws[f'U{row + OFFSET}'] = val
            row += 1

    col = 22 # start column
    for i in range(1, 10):
        # set market values found if they are given
        if f'gmvf{i}' in dict:
            row = row_start
            for val in dict[f'gmvf{i}']:
                ws.cell(row + OFFSET,col).value = val
                row += 1
        col += 1
        
        # set market value links found if they are given
        if f'gmvl{i}' in dict:
            row = row_start
            for val in dict[f'gmvl{i}']:
                ws.cell(row + OFFSET,col).value = val
                row += 1
        col += 1

    wb.save('Equipment New List.xlsx')

# sets given final prices in 'Equipment New List.xlsx'
def setExcel(dict):
    wb = pyxl.load_workbook('Equipment New List.xlsx')
    ws = wb.active
    n = len(dict['Search Terms'])

    # set column titles
    if 'Market Value Found' in dict:
        ws[f'Q1'] = 'Market Value Found'
    if 'Auction Value Found' in dict:
        ws[f'R1'] = 'Auction Value Found'
    if 'Auction Value Link' in dict:
        ws[f'S1'] = 'Auction Value Link'
    if 'Asking Value Found' in dict:
        ws[f'T1'] = 'Asking Value Found'
    if 'Asking Value Link' in dict:
        ws[f'U1'] = 'Asking Value Link'
    col = 22 # start column
    for i in range(1, 10):
        if f'gmvf{i}' in dict:
            ws.cell(1,col).value = f'General Market Value Found {i}'
        col += 1
        if f'gmvl{i}' in dict:
            ws.cell(1,col).value = f'General Market Value Link {i}'
        col += 1

    # set values
    row = 0 + OFFSET
    while row < n + OFFSET:
        ws[f'X{row}'] = dict['Search Terms'][row - OFFSET]
        if 'Market Value Found' in dict:
            ws[f'Q{row}'] = dict['Market Value Found'][row - OFFSET]
        if 'Auction Value Found' in dict:
            ws[f'R{row}'] = dict['Auction Value Found'][row - OFFSET]
        if 'Auction Value Link' in dict:
            ws[f'S{row}'] = dict['Auction Value Link'][row - OFFSET]
        if 'Asking Value Found' in dict:
            ws[f'T{row}'] = dict['Asking Value Found'][row - OFFSET]
        if 'Asking Value Link' in dict:
            ws[f'U{row}'] = dict['Asking Value Link'][row - OFFSET]
        col = 22 # start column
        for i in range(1, 10):
            if f'gmvf{i}' in dict:
                ws.cell(row,col).value = dict[f'gmvf{i}'][row - OFFSET]
            col += 1
            if f'gmvl{i}' in dict:
                ws.cell(row,col).value = dict[f'gmvl{i}'][row - OFFSET]
            col += 1
        row += 1

    wb.save('Equipment New List.xlsx')

def main():
    # output intro screen
    i = 10 
    print("")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
    print("           /$$$$$$                                          ")
    print("          /$$__  $$                                         ")
    print("         | $$  \__/  /$$$$$$   /$$$$$$   /$$$$$$   /$$$$$$$ ")
    print("         | $$       /$$__  $$ /$$__  $$ /$$__  $$ /$$_____/ ")
    print("         | $$      | $$$$$$$$| $$  \__/| $$$$$$$$|  $$$$$$  ")
    print("         | $$    $$| $$_____/| $$      | $$_____/ \____  $$ ")
    print("         |  $$$$$$/|  $$$$$$$| $$      |  $$$$$$$ /$$$$$$$/ ")
    print("          \______/  \_______/|__/       \_______/|_______/  \n\n")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
    print("             Heavy Pricing Tool developed by Xander Gardner\n\n")
    print("    - Please do NOT use computer while running\n")
    print("    - Expect to wait several hours\n")
    print("    - Refer to heavy_pricing_guide.pdf for details\n\n")
    print("Starting in 10 seconds")
    time.sleep(10)

    # record start time
    start_time_sec = time.time()

    # create a copy of the master
    original = r'Equipment Master List.xlsx'
    target = r'Equipment New List.xlsx'
    if not exists(target):
        shutil.copyfile(original, target)

    # get data from 'Equipment New List.xlsx'
    data = getExcelValues()

    # get online prices and create dictionary
    search_terms = get_search_terms(data)
    adv_search_terms = get_adv_search_terms(data)
    dict = getDict(data)
    dict['Search Terms'] = search_terms
    dict['Advanced Search Terms'] = adv_search_terms
    scrapeAskingValues(dict)
    scrapeAuctionValues(dict)
    scrapeGeneralMarketValues(dict)

    # calculate market price and validate values found
    success = setMarketValues(dict, data)
    
    # write data in dict to 'Equipment New List.xlsx'
    setExcel(dict)
    stop_time_sec = time.time()

    # output closing screen
    n = len(dict['Search Terms'])
    success_percent = round(success / n * 100, 2)
    total_time_sec = stop_time_sec - start_time_sec
    total_time_min = total_time_sec / 60
    time_hours = total_time_sec // 3600
    time_minutes = (total_time_sec % 3600) // 60
    time_seconds = total_time_sec % 60 // 1
    priced_per_minute = round(success / total_time_min, 2)

    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")
    print("COMPLETE!\n\n")
    print(f"    - {success} ({success_percent}%) market values found in {time_hours} hours {time_minutes} minutes and {time_seconds} seconds")
    print(f"    - {priced_per_minute} item market values per minute")
    print(f"    - Output is in \"Equipment New List.xlsx\"\n\n")
    print("Closing in 30 seconds")
    time.sleep(30)

if __name__ == "__main__":
    main()
